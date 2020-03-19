import socket
import openpyxl
import sys

print(sys.argv[1])

wb = openpyxl.load_workbook('Geçiş_Listeleri_V2.xlsx')
sheet = wb[sys.argv[1]]

for i in range(1, sheet.max_row+1):
    try:
        sheet.cell(row=i, column=2).value = socket.gethostbyname(sheet.cell(row=i, column=1).value)
        print(sheet.cell(row=i, column=1).value + "--->" + sheet.cell(row=i, column=2).value)
    except:
        sheet.cell(row=i, column=2).value = ""

wb.save('Geçiş_Listeleri_V2.xlsx')